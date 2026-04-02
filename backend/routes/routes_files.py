"""File upload and management routes for channels and tasks"""
import uuid
import os
import io
import shutil
import logging
from datetime import datetime, timezone
from pathlib import Path
from pydantic import BaseModel, Field
from fastapi import HTTPException, Request, UploadFile, File, Form
from typing import Optional, List

logger = logging.getLogger(__name__)

# File storage configuration
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", "/app/uploads"))
try:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
except OSError:
    logger.warning(f"Could not create upload directory {UPLOAD_DIR}")

# Allowed file types and max size
ALLOWED_EXTENSIONS = {
    # Images
    "png", "jpg", "jpeg", "gif", "webp", "svg", "ico",
    # Documents
    "pdf", "doc", "docx", "xls", "xlsx", "ppt", "pptx", "txt", "rtf", "odt",
    # Code files
    "py", "js", "jsx", "ts", "tsx", "html", "css", "scss", "json", "xml", "yaml", "yml",
    "java", "cpp", "c", "h", "hpp", "cs", "go", "rs", "rb", "php", "swift", "kt",
    "sql", "sh", "bash", "ps1", "bat", "cmd",
    # Data files
    "csv", "md", "markdown", "log", "env", "ini", "cfg", "conf",
    # Archives (view only, not extracted)
    "zip", "tar", "gz", "rar", "7z"
}

MAX_FILE_SIZE = 25 * 1024 * 1024  # 25MB

# MIME type mapping
MIME_TYPES = {
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "gif": "image/gif",
    "webp": "image/webp",
    "svg": "image/svg+xml",
    "pdf": "application/pdf",
    "doc": "application/msword",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "xls": "application/vnd.ms-excel",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "txt": "text/plain",
    "json": "application/json",
    "py": "text/x-python",
    "js": "text/javascript",
    "ts": "text/typescript",
    "html": "text/html",
    "css": "text/css",
    "md": "text/markdown",
    "csv": "text/csv",
    "xml": "application/xml",
    "zip": "application/zip",
}


def get_file_extension(filename: str) -> str:
    """Get file extension in lowercase"""
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def is_allowed_file(filename: str) -> bool:
    """Check if file type is allowed"""
    ext = get_file_extension(filename)
    return ext in ALLOWED_EXTENSIONS


def get_mime_type(filename: str) -> str:
    """Get MIME type for file"""
    ext = get_file_extension(filename)
    return MIME_TYPES.get(ext, "application/octet-stream")


def extract_document_text(content: bytes, filename: str, max_chars: int = 15000) -> str:
    """Extract readable text from supported document formats for AI consumption.
    Returns empty string for unsupported formats."""
    ext = get_file_extension(filename)

    try:
        # Plain text formats
        if ext in ("txt", "md", "markdown", "csv", "log", "json", "xml", "yaml", "yml",
                    "py", "js", "jsx", "ts", "tsx", "html", "css", "sql", "sh", "bash",
                    "java", "cpp", "c", "h", "go", "rs", "rb", "php", "swift", "kt",
                    "ini", "cfg", "conf", "env"):
            text = content.decode("utf-8", errors="replace")
            return text[:max_chars]

        # PDF
        if ext == "pdf":
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(content))
            pages = []
            for i, page in enumerate(reader.pages):
                page_text = page.extract_text() or ""
                if page_text.strip():
                    pages.append(f"[Page {i+1}]\n{page_text}")
                if sum(len(p) for p in pages) > max_chars:
                    break
            return "\n\n".join(pages)[:max_chars]

        # DOCX
        if ext == "docx":
            from docx import Document
            doc = Document(io.BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs)[:max_chars]

        # XLSX / CSV
        if ext in ("xlsx", "xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                rows = []
                for sheet in wb.sheetnames[:3]:
                    ws = wb[sheet]
                    rows.append(f"[Sheet: {sheet}]")
                    for row in ws.iter_rows(max_row=100, values_only=True):
                        row_str = " | ".join(str(c) if c is not None else "" for c in row)
                        if row_str.strip(" |"):
                            rows.append(row_str)
                return "\n".join(rows)[:max_chars]
            except Exception:
                return ""

    except Exception as e:
        logger.warning(f"Text extraction failed for {filename}: {e}")
    return ""


def register_file_routes(api_router, db, get_current_user, check_workspace_permission):

    async def _authed_user(request, workspace_id):
        user = await get_current_user(request)
        from nexus_utils import require_workspace_access
        await require_workspace_access(db, user, workspace_id)
        return user
    
    @api_router.post("/channels/{channel_id}/files")
    async def upload_channel_file(
        channel_id: str,
        request: Request,
        file: UploadFile = File(...),
        message: Optional[str] = Form(None)
    ):
        """Upload a file to a channel"""
        user = await get_current_user(request)
        
        # Get channel and workspace
        channel = await db.channels.find_one({"channel_id": channel_id})
        if not channel:
            raise HTTPException(404, "Channel not found")
        
        workspace_id = channel["workspace_id"]
        
        # Check permission
        await check_workspace_permission(db, workspace_id, user["user_id"], "upload_files")
        
        # Validate file
        if not file.filename:
            raise HTTPException(400, "No file provided")
        
        if not is_allowed_file(file.filename):
            raise HTTPException(400, f"File type not allowed. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}")
        
        # Check file size
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(400, f"File too large. Maximum size: {MAX_FILE_SIZE // (1024*1024)}MB")
        
        # Generate unique filename
        file_id = f"file_{uuid.uuid4().hex[:12]}"
        ext = get_file_extension(file.filename)
        stored_filename = f"{file_id}.{ext}"
        
        # Store file via storage abstraction
        from storage import save_file as _save_file
        await _save_file(workspace_id, stored_filename, content)
        
        # Create file record
        file_record = {
            "file_id": file_id,
            "workspace_id": workspace_id,
            "channel_id": channel_id,
            "task_session_id": None,
            "original_name": file.filename,
            "stored_name": stored_filename,
            "file_size": len(content),
            "mime_type": get_mime_type(file.filename),
            "extension": ext,
            "uploaded_by": user["user_id"],
            "uploader_name": user.get("name", "Unknown"),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Extract and store text for preview and AI context
        extracted_text = extract_document_text(content, file.filename)
        if extracted_text:
            file_record["extracted_text"] = extracted_text[:15000]
        
        await db.files.insert_one(file_record)
        
        # Create message with file attachment
        msg_content = message or f"Shared a file: {file.filename}"
        if extracted_text:
            # Include extracted text so AI agents can read the document
            truncated = extracted_text[:8000]
            msg_content += f"\n\n--- Document Content ({file.filename}) ---\n{truncated}"
            if len(extracted_text) > 8000:
                msg_content += f"\n[... truncated, {len(extracted_text)} total chars]"
            msg_content += "\n--- End Document ---"
        
        msg_id = f"msg_{uuid.uuid4().hex[:12]}"
        file_message = {
            "message_id": msg_id,
            "channel_id": channel_id,
            "sender_type": "human",
            "sender_id": user["user_id"],
            "sender_name": user.get("name", "Unknown"),
            "content": msg_content,
            "file_attachment": {
                "file_id": file_id,
                "name": file.filename,
                "size": len(content),
                "mime_type": get_mime_type(file.filename),
                "extension": ext,
                "has_extracted_text": bool(extracted_text),
                "text_length": len(extracted_text) if extracted_text else 0,
            },
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.messages.insert_one(file_message)
        
        file_record.pop("_id", None)
        file_record["has_extracted_text"] = bool(extracted_text)
        file_record["text_length"] = len(extracted_text) if extracted_text else 0
        return {
            "file": file_record,
            "message_id": msg_id,
            "download_url": f"/api/files/{file_id}/download"
        }
    
    @api_router.post("/task-sessions/{session_id}/files")
    async def upload_task_file(
        session_id: str,
        request: Request,
        file: UploadFile = File(...),
        message: Optional[str] = Form(None)
    ):
        """Upload a file to a task session"""
        user = await get_current_user(request)
        
        # Get task session
        session = await db.task_sessions.find_one({"session_id": session_id})
        if not session:
            raise HTTPException(404, "Task session not found")
        
        workspace_id = session["workspace_id"]
        
        # Check permission
        await check_workspace_permission(db, workspace_id, user["user_id"], "upload_files")
        
        # Validate file
        if not file.filename:
            raise HTTPException(400, "No file provided")
        
        if not is_allowed_file(file.filename):
            raise HTTPException(400, "File type not allowed")
        
        # Check file size
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(400, f"File too large. Maximum size: {MAX_FILE_SIZE // (1024*1024)}MB")
        
        # Generate unique filename
        file_id = f"file_{uuid.uuid4().hex[:12]}"
        ext = get_file_extension(file.filename)
        stored_filename = f"{file_id}.{ext}"
        
        # Create workspace directory
        ws_dir = UPLOAD_DIR / workspace_id
        ws_dir.mkdir(parents=True, exist_ok=True)
        
        # Save file
        file_path = ws_dir / stored_filename
        with open(file_path, "wb") as f:
            f.write(content)
        
        # Create file record
        file_record = {
            "file_id": file_id,
            "workspace_id": workspace_id,
            "channel_id": None,
            "task_session_id": session_id,
            "original_name": file.filename,
            "stored_name": stored_filename,
            "file_size": len(content),
            "mime_type": get_mime_type(file.filename),
            "extension": ext,
            "uploaded_by": user["user_id"],
            "uploader_name": user.get("name", "Unknown"),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.files.insert_one(file_record)
        
        # Create task message with file attachment
        msg_content = message or f"Shared a file: {file.filename}"
        msg_id = f"tsm_{uuid.uuid4().hex[:12]}"
        file_message = {
            "message_id": msg_id,
            "session_id": session_id,
            "sender_type": "human",
            "sender_id": user["user_id"],
            "sender_name": user.get("name", "Unknown"),
            "content": msg_content,
            "file_attachment": {
                "file_id": file_id,
                "name": file.filename,
                "size": len(content),
                "mime_type": get_mime_type(file.filename),
                "extension": ext
            },
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.task_session_messages.insert_one(file_message)
        
        # Update message count
        await db.task_sessions.update_one(
            {"session_id": session_id},
            {"$inc": {"message_count": 1}}
        )
        
        file_record.pop("_id", None)
        return {
            "file": file_record,
            "message_id": msg_id,
            "download_url": f"/api/files/{file_id}/download"
        }
    
    @api_router.get("/files/{file_id}")
    async def get_file_info(file_id: str, request: Request):
        """Get file metadata"""
        user = await get_current_user(request)
        
        file_record = await db.files.find_one({"file_id": file_id}, {"_id": 0})
        if not file_record:
            raise HTTPException(404, "File not found")
        
        # Check workspace access
        await check_workspace_permission(db, file_record["workspace_id"], user["user_id"], "view_workspace")
        
        return file_record
    
    @api_router.get("/files/{file_id}/download")
    async def download_file(file_id: str, request: Request):
        """Download a file"""
        user = await get_current_user(request)
        
        file_record = await db.files.find_one({"file_id": file_id})
        if not file_record:
            raise HTTPException(404, "File not found")
        
        # Check workspace access
        await check_workspace_permission(db, file_record["workspace_id"], user["user_id"], "view_workspace")
        
        # Read from storage abstraction
        from storage import get_file as _get_file, get_file_url as _get_file_url, STORAGE_BACKEND
        if STORAGE_BACKEND == "s3":
            url = await _get_file_url(file_record["workspace_id"], file_record["stored_name"])
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url=url)
        
        file_bytes = await _get_file(file_record["workspace_id"], file_record["stored_name"])
        if not file_bytes:
            raise HTTPException(404, "File data not found")
        
        from nexus_utils import sanitize_filename
        safe_name = sanitize_filename(file_record.get("original_name", "download"))
        from fastapi.responses import Response as RawResp
        return RawResp(
            content=file_bytes,
            media_type=file_record["mime_type"],
            headers={"Content-Disposition": f'attachment; filename="{safe_name}"'}
        )
    
    @api_router.get("/files/{file_id}/text")
    async def get_file_text(file_id: str, request: Request):
        """Get extracted text from a file (for DOCX, PDF, etc.)."""
        user = await get_current_user(request)
        file_record = await db.files.find_one({"file_id": file_id}, {"_id": 0})
        if not file_record:
            raise HTTPException(404, "File not found")
        await check_workspace_permission(db, file_record["workspace_id"], user["user_id"], "view_workspace")
        text = file_record.get("extracted_text", "")
        if not text:
            return {"text": "", "message": "No extracted text available for this file type"}
        return {"text": text, "file_id": file_id, "filename": file_record.get("original_name", "")}

    @api_router.get("/files/{file_id}/preview")
    async def preview_file(file_id: str, request: Request):
        """Serve file for inline preview (Content-Disposition: inline)."""
        user = await get_current_user(request)
        file_record = await db.files.find_one({"file_id": file_id})
        if not file_record:
            raise HTTPException(404, "File not found")
        await check_workspace_permission(db, file_record["workspace_id"], user["user_id"], "view_workspace")
        from storage import get_file as _get_file
        file_bytes = await _get_file(file_record["workspace_id"], file_record["stored_name"])
        if not file_bytes:
            raise HTTPException(404, "File data not found")
        from fastapi.responses import Response as RawResp
        from nexus_utils import sanitize_filename
        safe_name = sanitize_filename(file_record.get("original_name", "download"))
        return RawResp(
            content=file_bytes,
            media_type=file_record["mime_type"],
            headers={"Content-Disposition": f'inline; filename="{safe_name}"'}
        )

    @api_router.delete("/files/{file_id}")
    async def delete_file(file_id: str, request: Request):

        """Delete a file"""
        user = await get_current_user(request)
        
        file_record = await db.files.find_one({"file_id": file_id})
        if not file_record:
            raise HTTPException(404, "File not found")
        
        # Check permission (uploader or admin can delete)
        role = await check_workspace_permission(db, file_record["workspace_id"], user["user_id"], "delete_files")
        
        if file_record["uploaded_by"] != user["user_id"] and role != "admin":
            raise HTTPException(403, "Only the uploader or admin can delete this file")
        
        # Delete from storage
        from storage import delete_file as _delete_file
        await _delete_file(file_record["workspace_id"], file_record["stored_name"])
        
        # Delete record
        await db.files.delete_one({"file_id": file_id})
        
        return {"status": "deleted"}
    
    @api_router.get("/channels/{channel_id}/files")
    async def list_channel_files(channel_id: str, request: Request):
        """List all files in a channel"""
        user = await get_current_user(request)
        
        channel = await db.channels.find_one({"channel_id": channel_id})
        if not channel:
            raise HTTPException(404, "Channel not found")
        
        # Check workspace access
        await check_workspace_permission(db, channel["workspace_id"], user["user_id"], "view_workspace")
        
        files = await db.files.find(
            {"channel_id": channel_id},
            {"_id": 0}
        ).sort("created_at", -1).to_list(100)
        
        return {"files": files, "count": len(files)}
    
    @api_router.get("/task-sessions/{session_id}/files")
    async def list_task_files(session_id: str, request: Request):
        """List all files in a task session"""
        user = await get_current_user(request)
        
        session = await db.task_sessions.find_one({"session_id": session_id})
        if not session:
            raise HTTPException(404, "Task session not found")
        
        # Check workspace access
        await check_workspace_permission(db, session["workspace_id"], user["user_id"], "view_workspace")
        
        files = await db.files.find(
            {"task_session_id": session_id},
            {"_id": 0}
        ).sort("created_at", -1).to_list(100)
        
        return {"files": files, "count": len(files)}
    
    @api_router.get("/workspaces/{workspace_id}/files")
    async def list_workspace_files(workspace_id: str, request: Request):
        """List all files in a workspace"""
        user = await _authed_user(request, workspace_id)
        
        # Check workspace access
        await check_workspace_permission(db, workspace_id, user["user_id"], "view_workspace")
        """List all files in a workspace"""
        user = await _authed_user(request, workspace_id)
        
        # Check workspace access
        await check_workspace_permission(db, workspace_id, user["user_id"], "view_workspace")
        
        files = await db.files.find(
            {"workspace_id": workspace_id},
            {"_id": 0}
        ).sort("created_at", -1).to_list(500)
        
        # Calculate total size
        total_size = sum(f.get("file_size", 0) for f in files)
        
        return {
            "files": files,
            "count": len(files),
            "total_size": total_size,
            "total_size_mb": round(total_size / (1024 * 1024), 2)
        }
